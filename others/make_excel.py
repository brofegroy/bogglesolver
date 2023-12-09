import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

board = ["OFEQ", "HATE", "VSRK", "WNIM"]#OFEQ HATE VSRK WNIM
formable_words = [["tea","tae","eat","eta"],["sean","four"]] + [[] for _ in range(20)]

def make_excel(board,formable_words):
    print(board)
    # current_directory = os.getcwd()
    # save_path = os.path.join(current_directory, '..', 'solution.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    #setting column width
    for col_index in range(3, 17):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index-2)].width = 18

    #board
    for row_index, row in enumerate(board, start=3):
        for col_index, char in enumerate(row, start=6):
            if char == "Q":
                char = "Qu"
            cell = sheet.cell(row=row_index, column=col_index, value=char)
            cell.font = Font(size=30)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    #board border
    for col_index in range(6, 10):
        for row_index in range(3, 7):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.border = Border(
                left=Side(style='double'),
                right=Side(style='double'),
                top=Side(style='double'),
                bottom=Side(style='double')
            )

    #set board height
    default_row_height = 60
    for row_index in range(3, 7):
        sheet.row_dimensions[row_index].height = default_row_height

    #annotationg the columns
    for col_index in range(3, 17):
        cell = sheet.cell(row=10, column=col_index-2, value=f"count: {len(formable_words[col_index-3])}")
        cell.font = Font(size=11)

    sheet['I7'] = f"total words: {sum(len(sublist) for sublist in formable_words)}"

    for col_index in range(3, 17):
        cell = sheet.cell(row=11, column=col_index-2, value=f"{col_index} letter words")
        cell.font = Font(size=14)
        cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    #display the words
    for col_index, data_list in enumerate(formable_words, start=1):
        # Write each element of the inner list to the corresponding column
        for row_index, value in enumerate(data_list, start=12):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.font = Font(size=14)

    # Save the workbook
    workbook.save('solution.xlsx')
    return